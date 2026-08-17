"""Build and send the morning digest: an Excel workbook plus an HTML email.

Two sections, because they answer different questions. *New since last digest*
is what you act on today. *Standing shortlist* is everything still eligible you
have not marked applied or skipped, so nothing quietly falls off the bottom
while you were busy.
"""

from __future__ import annotations

import json
import sqlite3
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import comp as comp_mod
from . import coverage as cov, db, rank, track

COLUMNS = [
    ("Score", 8),
    ("Track", 9),
    ("Company", 18),
    ("Title", 46),
    ("Location", 24),
    ("Posted", 12),
    ("Age", 6),
    ("Salary", 16),
    ("Why it matched", 60),
    ("Gaps vs JD", 34),
    ("Est. comp", 30),
    ("Apply", 12),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_LINK_FONT = Font(color="2563EB", underline="single")


@dataclass(slots=True)
class Digest:
    day: str
    new: list[sqlite3.Row]
    standing: list[sqlite3.Row]
    stretch: list[sqlite3.Row] = field(default_factory=list)
    follow_ups: list = field(default_factory=list)
    pipeline: dict = field(default_factory=dict)
    xlsx: Path | None = None


# ------------------------------------------------------------------ assemble


def _age_days(posted_at: str | None) -> int | None:
    if not posted_at:
        return None
    try:
        return (datetime.now(UTC) - datetime.fromisoformat(posted_at)).days
    except ValueError:
        return None


def _salary(row: sqlite3.Row) -> str:
    lo, hi, ccy = row["salary_min"], row["salary_max"], row["salary_ccy"] or ""
    if not lo:
        return ""
    fmt = lambda v: f"{v/1000:.0f}K" if v < 1_000_000 else f"{v/1_000_000:.1f}M"
    return f"{ccy} {fmt(lo)}–{fmt(hi)}".strip() if hi else f"{ccy} {fmt(lo)}+".strip()


def _comp_text(row, cache, floor) -> str:
    entry = comp_mod.lookup(cache, row["company"], row["title"])
    return entry.summary(floor) if entry and entry.found else ""


def _reasons(row: sqlite3.Row, limit: int = 4) -> str:
    try:
        return ", ".join(json.loads(row["reasons"] or "[]")[:limit])
    except (json.JSONDecodeError, TypeError):
        return ""


def build(conn: sqlite3.Connection, *, day: str | None = None, standing_limit: int = 30) -> Digest:
    day = day or date.today().isoformat()
    eligible = rank.shortlist(conn, limit=10_000)

    already = db.load_seen()
    states = db.load_status()
    closed = {jid for jid, s in states.items() if s.get("state") in {"applied", "skipped"}}

    new = [r for r in eligible if r["id"] not in already and r["id"] not in closed]
    standing = [r for r in eligible if r["id"] not in closed][:standing_limit]

    # Near-misses on seniority, kept in their own section so "SDE II" never
    # gets silently mixed in with roles that clear every bar.
    stretch = [
        r for r in rank.shortlist(conn, limit=standing_limit, tier="stretch")
        if r["id"] not in closed
    ]
    apps = track.load(conn)
    return Digest(
        day=day, new=new, standing=standing, stretch=stretch,
        follow_ups=track.follow_ups(apps),
        pipeline=track.summary(apps) | {
            "applied_this_week": track.applied_this_week(apps),
            "response_rate": track.response_rate(apps),
        },
    )


# --------------------------------------------------------------------- excel


def _write_sheet(ws, rows, owned=None, comp_cache=None, floor=None) -> None:
    owned = owned if owned is not None else set()
    comp_cache = comp_cache if comp_cache is not None else {}
    ws.append([name for name, _ in COLUMNS])
    for idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        cell = ws.cell(row=1, column=idx)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        age = _age_days(row["posted_at"])
        ws.append([
            round(row["score"], 1),
            row["profile"],
            row["company"],
            row["title"],
            row["location"] or "",
            (row["posted_at"] or "")[:10],
            age if age is not None else "",
            _salary(row),
            _reasons(row),
            cov.analyse(row["description"], owned).summary(4),
            _comp_text(row, comp_cache, floor),
            "Open",
        ])
        # Hyperlink the last cell rather than pasting a raw URL into a column:
        # apply links routinely run past 120 characters.
        link = ws.cell(row=ws.max_row, column=len(COLUMNS))
        link.hyperlink, link.font = row["url"], _LINK_FONT

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def write_xlsx(digest: Digest, directory: Path | None = None) -> Path:
    directory = directory or db.DIGEST_DIR
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"autowork-{digest.day}.xlsx"

    cfg = rank.load_config()
    owned = cov.candidate_terms(cfg)
    cache = comp_mod.load()
    floor = cfg["candidate"].get("current_ctc_lpa")
    wb = Workbook()
    _write_sheet(wb.active, digest.new, owned, cache, floor)
    wb.active.title = f"New ({len(digest.new)})"
    _write_sheet(wb.create_sheet(f"Shortlist ({len(digest.standing)})"), digest.standing, owned, cache, floor)
    _write_sheet(wb.create_sheet(f"Stretch ({len(digest.stretch)})"), digest.stretch, owned, cache, floor)
    wb.save(path)
    digest.xlsx = path
    return path


# --------------------------------------------------------------------- email


def _esc(text: str) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
    )


def _comp_html(row, cache: dict, floor: float | None) -> str:
    entry = comp_mod.lookup(cache, row["company"], row["title"])
    if not (entry and entry.found):
        return ""
    below = floor and entry.median and entry.median < floor
    colour = "#b91c1c" if below else "#047857"
    return (
        f"<br><span style='color:{colour};font-size:12px;font-weight:600'>"
        f"{_esc(entry.summary(floor))}</span>"
    )


def _follow_up_html(digest: Digest) -> str:
    """Placed above the new roles: it is the only part of the digest with a
    deadline. New postings keep; a follow-up window closes."""
    if not digest.follow_ups:
        return ""
    items = "".join(
        f"<li style='margin:4px 0'><a href='{_esc(a.url)}' style='color:#92400e;"
        f"font-weight:600;text-decoration:none'>{_esc(a.title or a.job_id)}</a>"
        f" — {_esc(a.company)} · applied {a.days_since}d ago</li>"
        for a in digest.follow_ups
    )
    return (
        "<div style='background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;"
        "padding:12px 16px;margin:0 0 20px'>"
        f"<strong style='color:#92400e;font-size:14px'>Follow up "
        f"({len(digest.follow_ups)})</strong>"
        f"<ul style='margin:8px 0 0;padding-left:18px;color:#78350f;font-size:13px'>{items}</ul>"
        "</div>"
    )


def _pipeline_html(digest: Digest) -> str:
    p = digest.pipeline
    if not p or not p.get("live"):
        return ""
    bits = [f"{p['live']} live", f"{p.get('applied_this_week', 0)} applied this week"]
    if p.get("response_rate") is not None:
        bits.append(f"{p['response_rate']:.0%} response rate")
    if p.get("cold"):
        bits.append(f"{p['cold']} gone cold")
    return (
        "<p style='margin:0 0 16px;color:#6b7280;font-size:12px'>"
        + " · ".join(bits) + "</p>"
    )


def _rows_html(rows, cache: dict | None = None, floor: float | None = None) -> str:
    cache = cache if cache is not None else {}
    if not rows:
        return "<p style='color:#6b7280;margin:8px 0 20px'>Nothing new today.</p>"
    cells = []
    for row in rows:
        age = _age_days(row["posted_at"])
        age_txt = f"{age}d ago" if age is not None else ""
        cells.append(
            "<tr>"
            f"<td style='padding:10px 8px;border-bottom:1px solid #e5e7eb;font-weight:600;"
            f"color:#111827;white-space:nowrap'>{row['score']:.0f}</td>"
            f"<td style='padding:10px 8px;border-bottom:1px solid #e5e7eb'>"
            f"<a href='{_esc(row['url'])}' style='color:#1d4ed8;text-decoration:none;"
            f"font-weight:600'>{_esc(row['title'])}</a><br>"
            f"<span style='color:#374151;font-size:13px'>{_esc(row['company'])}</span>"
            f"<span style='color:#9ca3af;font-size:13px'> · {_esc(row['location'] or '—')}"
            f" · {age_txt}</span><br>"
            f"<span style='color:#6b7280;font-size:12px'>{_esc(_reasons(row, 3))}</span>"
            f"{_comp_html(row, cache, floor)}"
            "</td>"
            f"<td style='padding:10px 8px;border-bottom:1px solid #e5e7eb;color:#6b7280;"
            f"font-size:12px;white-space:nowrap'>{_esc(row['profile'])}</td>"
            "</tr>"
        )
    return (
        "<table style='width:100%;border-collapse:collapse;font-family:"
        "-apple-system,Segoe UI,Roboto,sans-serif;font-size:14px'>"
        + "".join(cells)
        + "</table>"
    )


def render_html(digest: Digest) -> str:
    cfg = rank.load_config()
    cache = comp_mod.load()
    floor = cfg["candidate"].get("current_ctc_lpa")
    return f"""<body style="margin:0;padding:24px;background:#f9fafb">
<div style="max-width:720px;margin:0 auto;background:#fff;border:1px solid #e5e7eb;
            border-radius:10px;padding:24px;font-family:-apple-system,Segoe UI,Roboto,sans-serif">
  <h2 style="margin:0 0 4px;font-size:18px;color:#111827">AutoWork · {digest.day}</h2>
  <p style="margin:0 0 20px;color:#6b7280;font-size:13px">
    {len(digest.new)} new · {len(digest.standing)} shortlisted · {len(digest.stretch)} stretch
  </p>
  {_pipeline_html(digest)}
  {_follow_up_html(digest)}
  <h3 style="margin:0 0 4px;font-size:14px;color:#111827">New since last digest</h3>
  {_rows_html(digest.new, cache, floor)}
  <h3 style="margin:24px 0 4px;font-size:14px;color:#111827">Standing shortlist</h3>
  {_rows_html(digest.standing[:12], cache, floor)}
  <h3 style="margin:24px 0 4px;font-size:14px;color:#111827">
    Stretch <span style="font-weight:400;color:#9ca3af;font-size:12px">
    · SDE II or a 3-year ask — a reach, not out of reach</span></h3>
  {_rows_html(digest.stretch[:8], cache, floor)}
  <p style="margin:24px 0 0;color:#9ca3af;font-size:12px;border-top:1px solid #e5e7eb;
            padding-top:12px">
    Full ranked list in the attached workbook.
  </p>
</div></body>"""


# Sending lives in `deliver.py` — this module only builds.
